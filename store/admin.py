from django.contrib import admin
from django.utils import timezone
from django.utils.html import format_html
from django import forms
from datetime import timedelta
from django.contrib import messages
from django.urls import path, reverse
from django.shortcuts import render, redirect
from django.db import transaction
from django.db.models import Sum, Case, When, Value, IntegerField, Count
from django.db.models.functions import Trim
from decimal import Decimal, InvalidOperation
import re
import zipfile
from django.contrib.admin.helpers import ActionForm
from django.http import HttpResponse
import csv
from io import StringIO

from .models import (
    Book,
    BookImage,
    Cart,
    CartItem,
    CartBundleItem,
    Order,
    OrderItem,
    Wishlist,
    Category,
    Subject,
    SyllabusPDF,
    Review,
    Coupon,
    GalleryItem,
    UserProfile,
    UserAddress,
    Bundle,
    PublishWithUsSubmission,
    Banner,
    SearchQueryLog,
    DictionaryQueryLog,
    DictionaryTermOpenLog,
    UnaniReferenceSource,
    AuditLog,
    SiteSettings,
    IKSCoinsSettings,
    IKSWallet,
    IKSWalletTransaction,
    ClassicalWeightUnit,
    UnaniTerm,
)
from .admin_site import IdaraAdminSite
from .coins import manual_adjust_wallet, queue_order_pending_rewards, process_due_pending_rewards_for_user
from .dictionary_text import (
    normalize_latin_search_text,
    normalize_script_text,
    normalize_transliteration_text,
    transliteration_invalid_chars,
    validate_transliteration_style,
)

admin_site = IdaraAdminSite(name="idara_admin")

def _log_audit(request, action, obj, changes):
    AuditLog.objects.create(
        user=request.user if request and request.user.is_authenticated else None,
        action=action,
        model_name=obj.__class__.__name__,
        object_id=obj.pk or 0,
        object_repr=str(obj)[:200],
        changes=changes,
    )


def _normalize_header(value):
    return str(value or "").strip().lower().replace("_", " ")


def _split_multi(value):
    if value in (None, ""):
        return []
    return [part.strip() for part in re.split(r"[\/,|]+", str(value)) if part.strip()]


def _coerce_bool(value, default=None):
    if value in (None, ""):
        return default
    val = str(value).strip().lower()
    if val in {"1", "true", "yes", "y", "on"}:
        return True
    if val in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _coerce_int(value, default=None):
    if value in (None, ""):
        return default
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _coerce_decimal(value, default=None):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return default


def _load_tabular_rows(uploaded_file, expected_headers=None, overflow_merge_header=None):
    name = (getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith(".csv"):
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        if isinstance(raw, str):
            csv_text = raw
        else:
            csv_text = None
            for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"):
                try:
                    csv_text = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if csv_text is None:
                raise ValueError("Could not decode CSV. Save the file as UTF-8 CSV or upload XLSX and try again.")
        reader = csv.reader(StringIO(csv_text))
        rows = list(reader)
    elif name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("openpyxl is not installed. Add it to requirements and deploy.") from exc
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    else:
        raise ValueError("Unsupported file type. Upload CSV or XLSX.")

    if not rows:
        return []

    normalized_expected = [_normalize_header(item) for item in (expected_headers or ()) if _normalize_header(item)]
    first_row_headers = [_normalize_header(item) for item in rows[0]]
    has_header_row = True
    if normalized_expected:
        has_header_row = any(header in normalized_expected for header in first_row_headers if header)

    headers = first_row_headers if has_header_row else normalized_expected
    data_rows = rows[1:] if has_header_row else rows
    if not headers:
        return []

    merge_index = None
    normalized_merge_header = _normalize_header(overflow_merge_header) if overflow_merge_header else None
    if normalized_merge_header and normalized_merge_header in headers:
        merge_index = headers.index(normalized_merge_header)

    parsed_rows = []
    row_start = 2 if has_header_row else 1
    for index, raw_row in enumerate(data_rows, start=row_start):
        row_values = list(raw_row)
        if merge_index is not None and len(row_values) > len(headers):
            overflow_count = len(row_values) - len(headers)
            merged_chunk = row_values[merge_index:merge_index + overflow_count + 1]
            merged_value = ",".join("" if value is None else str(value) for value in merged_chunk)
            row_values = (
                row_values[:merge_index]
                + [merged_value]
                + row_values[merge_index + overflow_count + 1:]
            )

        row_dict = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            value = row_values[col_index] if col_index < len(row_values) else None
            row_dict[header] = value
        if any(value not in (None, "") for value in row_dict.values()):
            row_dict["_row_number"] = index
            parsed_rows.append(row_dict)
    return parsed_rows


class BulkImportForm(forms.Form):
    file = forms.FileField(help_text="Upload CSV or XLSX file.")
    dry_run = forms.BooleanField(required=False, initial=False, help_text="Validate only.")


class BulkImportAdminMixin:
    bulk_import_template = "admin/store/common/import_data.html"
    bulk_import_changelist_template = "admin/store/common/change_list_with_import.html"
    bulk_import_title = "Bulk Import"
    bulk_import_help = ""
    bulk_import_columns = ()
    bulk_import_sample_rows = ()
    bulk_import_overflow_merge_column = None

    def get_bulk_import_url_name(self):
        return f"{self.model._meta.app_label}_{self.model._meta.model_name}_import_data"

    def get_bulk_import_sample_url_name(self):
        return f"{self.model._meta.app_label}_{self.model._meta.model_name}_sample_csv"

    def get_bulk_export_url_name(self):
        return f"{self.model._meta.app_label}_{self.model._meta.model_name}_export_data"

    def get_bulk_export_xlsx_url_name(self):
        return f"{self.model._meta.app_label}_{self.model._meta.model_name}_export_data_xlsx"

    def get_bulk_import_sample_filename(self):
        return f"{self.model._meta.model_name}_sample.csv"

    def get_bulk_export_filename(self):
        return f"{self.model._meta.model_name}_export.csv"

    def get_bulk_export_xlsx_filename(self):
        return f"{self.model._meta.model_name}_export.xlsx"

    def get_bulk_import_sample_rows(self):
        return self.bulk_import_sample_rows

    def get_bulk_export_queryset(self, request):
        return self.model.objects.all().order_by("id")

    def _default_export_value(self, obj, field_name):
        if not hasattr(obj, field_name):
            return ""
        value = getattr(obj, field_name)
        if callable(value):
            value = value()
        if isinstance(value, bool):
            return "true" if value else "false"
        if value is None:
            return ""
        return str(value)

    def get_bulk_export_row(self, obj, columns):
        row = []
        for col in columns:
            field_name = str(col or "").strip().lower().replace(" ", "_")
            row.append(self._default_export_value(obj, field_name))
        return row

    def get_bulk_import_rows(self, request):
        if "file" not in request.FILES:
            return []
        return _load_tabular_rows(
            request.FILES["file"],
            expected_headers=self.bulk_import_columns,
            overflow_merge_header=self.bulk_import_overflow_merge_column,
        )

    def import_row(self, row_data):
        raise NotImplementedError("import_row must be implemented.")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-data/",
                self.admin_site.admin_view(self.import_data_view),
                name=self.get_bulk_import_url_name(),
            ),
            path(
                "import-data/sample-csv/",
                self.admin_site.admin_view(self.download_sample_csv_view),
                name=self.get_bulk_import_sample_url_name(),
            ),
            path(
                "export-data/",
                self.admin_site.admin_view(self.export_data_view),
                name=self.get_bulk_export_url_name(),
            ),
            path(
                "export-data-xlsx/",
                self.admin_site.admin_view(self.export_data_xlsx_view),
                name=self.get_bulk_export_xlsx_url_name(),
            ),
        ]
        return custom_urls + urls

    def download_sample_csv_view(self, request):
        columns = list(self.bulk_import_columns or ())
        if not columns:
            messages.error(request, "No sample columns configured for this importer.")
            return render(request, self.bulk_import_template, {
                "form": BulkImportForm(),
                "title": self.bulk_import_title,
                "bulk_import_help": self.bulk_import_help,
                "bulk_import_columns": self.bulk_import_columns,
                "sample_csv_url": reverse(f"admin:{self.get_bulk_import_sample_url_name()}"),
                "changelist_url": reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                ),
            })

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.get_bulk_import_sample_filename()}"'
        writer = csv.writer(response)
        writer.writerow(columns)

        for row in self.get_bulk_import_sample_rows() or ():
            if isinstance(row, dict):
                writer.writerow([row.get(col, "") for col in columns])
            elif isinstance(row, (list, tuple)):
                values = list(row[:len(columns)])
                if len(values) < len(columns):
                    values += [""] * (len(columns) - len(values))
                writer.writerow(values)
            else:
                writer.writerow([""] * len(columns))

        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["import_button_label"] = "Bulk Import"
        extra_context["import_button_url"] = reverse(f"admin:{self.get_bulk_import_url_name()}")
        extra_context["export_button_label"] = "Export All CSV"
        extra_context["export_button_url"] = reverse(f"admin:{self.get_bulk_export_url_name()}")
        extra_context["export_xlsx_button_label"] = "Export All XLSX"
        extra_context["export_xlsx_button_url"] = reverse(f"admin:{self.get_bulk_export_xlsx_url_name()}")
        return super().changelist_view(request, extra_context=extra_context)

    def export_data_view(self, request):
        columns = list(self.bulk_import_columns or ())
        if not columns:
            messages.error(request, "No export columns configured for this section.")
            return redirect(
                reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.get_bulk_export_filename()}"'
        writer = csv.writer(response)
        writer.writerow(columns)
        for obj in self.get_bulk_export_queryset(request):
            writer.writerow(self.get_bulk_export_row(obj, columns))
        return response

    def export_data_xlsx_view(self, request):
        columns = list(self.bulk_import_columns or ())
        if not columns:
            messages.error(request, "No export columns configured for this section.")
            return redirect(
                reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )
            )
        try:
            from openpyxl import Workbook
        except Exception:
            messages.error(request, "openpyxl is not installed on this server.")
            return redirect(
                reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        ws.append(columns)
        for obj in self.get_bulk_export_queryset(request):
            ws.append(self.get_bulk_export_row(obj, columns))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.get_bulk_export_xlsx_filename()}"'
        wb.save(response)
        return response

    def import_data_view(self, request):
        form = BulkImportForm(request.POST or None, request.FILES or None)
        if request.method == "POST" and form.is_valid():
            dry_run = form.cleaned_data["dry_run"]
            created = 0
            updated = 0
            skipped = 0
            errors = []
            try:
                rows = self.get_bulk_import_rows(request)
            except Exception as exc:
                messages.error(request, str(exc))
                rows = None

            if rows is not None:
                with transaction.atomic():
                    for row in rows:
                        try:
                            status = self.import_row(row)
                        except Exception as exc:
                            errors.append(f"Row {row.get('_row_number', '?')}: {exc}")
                            continue
                        if status == "created":
                            created += 1
                        elif status == "updated":
                            updated += 1
                        else:
                            skipped += 1

                    if dry_run:
                        transaction.set_rollback(True)

            if rows is not None:
                summary = (
                    f"Import completed. Created: {created}, Updated: {updated}, "
                    f"Skipped: {skipped}, Errors: {len(errors)}"
                )
                if dry_run:
                    summary += " (dry run)"
                messages.success(request, summary)
                for item in errors[:20]:
                    messages.error(request, item)
                if len(errors) > 20:
                    messages.error(request, f"Additional errors not shown: {len(errors) - 20}")

        context = {
            "form": form,
            "title": self.bulk_import_title,
            "bulk_import_help": self.bulk_import_help,
            "bulk_import_columns": self.bulk_import_columns,
            "sample_csv_url": reverse(f"admin:{self.get_bulk_import_sample_url_name()}"),
            "changelist_url": reverse(
                f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
            ),
        }
        return render(request, self.bulk_import_template, context)


class ProductivityAdminMixin:
    save_on_top = True
    list_per_page = 50
    list_max_show_all = 200
    preserve_filters = True
    show_full_result_count = False
    actions_on_top = True
    actions_on_bottom = True

# ======================
# CATEGORY ADMIN
# ======================

@admin.register(Category)
class CategoryAdmin(BulkImportAdminMixin, ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    search_fields = ("name",)
    ordering = ("name",)
    change_list_template = BulkImportAdminMixin.bulk_import_changelist_template
    bulk_import_title = "Bulk Import Categories"
    bulk_import_help = "Required: name. Optional: slug."
    bulk_import_columns = ("name", "slug")
    bulk_import_sample_rows = (
        {"name": "Unani Classics", "slug": "unani-classics"},
        {"name": "Pharmacology", "slug": "pharmacology"},
    )

    def import_row(self, row_data):
        name = str(row_data.get("name") or "").strip()
        if not name:
            raise ValueError("name is required.")

        slug = str(row_data.get("slug") or "").strip()
        category, created = Category.objects.get_or_create(name=name)
        if slug:
            category.slug = slug
            category.save()
            return "created" if created else "updated"
        return "created" if created else "updated"


# ======================
# BOOK IMAGE INLINE
# ======================

class BookImageInline(admin.TabularInline):
    model = BookImage
    extra = 1
    fields = ("image", "is_watermarked")
    readonly_fields = ("is_watermarked",)
    show_change_link = True


class BundleBookInline(admin.TabularInline):
    model = Bundle.books.through
    extra = 1


@admin.register(Bundle)
class BundleAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("name", "bundle_price", "is_active", "created_at")
    list_filter = ("is_active", "created_at")
    search_fields = ("name",)
    inlines = [BundleBookInline]
    exclude = ("books",)


# ======================
# BOOK ADMIN
# ======================

@admin.register(Book)
class BookAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "system_id",
        "title",
        "author",
        "price",
        "mrp_price",
        "discount_display",
        "stock",
        "is_active",
        "category",
        "is_bestseller",
        "is_trending",
        "is_new_arrival",
        "is_featured",
    )

    list_filter = (
        "category",
        ("subjects", admin.RelatedOnlyFieldListFilter),
        "is_active",
        "is_bestseller",
        "is_trending",
        "is_new_arrival",
        "is_featured",
    )
    search_fields = ("system_id", "isbn", "title", "author", "description")
    list_select_related = ("category",)
    autocomplete_fields = ("category", "subjects")

    list_editable = (
        "price",
        "mrp_price",
        "stock",
        "is_active",
        "is_bestseller",
        "is_trending",
        "is_new_arrival",
        "is_featured",
    )

    readonly_fields = ("discount_display", "main_cover_preview")
    inlines = [BookImageInline]

    change_list_template = "admin/store/book/change_list.html"

    fieldsets = (
        ("Basic Information", {
            "fields": ("system_id", "category", "subjects", "title", "author", "description")
        }),
        ("Specifications", {
            "fields": ("isbn", "published_year", "binding", "pages", "weight")
        }),
        ("Pricing", {
            "fields": ("price", "mrp_price", "discount_display")
        }),
        ("Stock", {
            "fields": ("stock", "is_active")
        }),
        ("Homepage Flags", {
            "fields": ("is_bestseller", "is_trending", "is_new_arrival", "is_featured")
        }),
        ("Images", {
            "fields": ("main_cover_preview", "main_cover")
        }),
        ("Book Files", {
            "fields": ("toc_pdf", "sample_pdf")
        }),
    )

    ordering = ("title",)

    def main_cover_preview(self, obj):
        if obj and obj.main_cover:
            return format_html(
                '<a href="{0}" target="_blank" rel="noopener noreferrer">'
                '<img src="{0}" alt="{1}" style="max-height:120px;border:1px solid #ddd;border-radius:4px;" />'
                "</a>",
                obj.main_cover.url,
                obj.title or "Book cover",
            )
        return "No cover uploaded"
    main_cover_preview.short_description = "Current cover"

    class PriceUpdateActionForm(ActionForm):
        percentage = forms.DecimalField(
            required=False,
            label="Price %",
            help_text="Use positive or negative (e.g., 10 or -5).",
        )
        apply_to_mrp = forms.BooleanField(
            required=False,
            initial=True,
            label="Apply to MRP too",
        )

    action_form = PriceUpdateActionForm
    actions = (
        "mark_as_featured",
        "remove_from_featured",
        "mark_as_new_arrival",
        "remove_from_new_arrival",
        "mark_as_bestseller",
        "remove_from_bestseller",
        "mark_as_trending",
        "remove_from_trending",
        "activate_books",
        "deactivate_books",
        "bulk_update_price",
        "export_books_csv",
    )

    class ImportBooksForm(forms.Form):
        file = forms.FileField(help_text="Upload CSV or XLSX file.")
        dry_run = forms.BooleanField(required=False, initial=False, help_text="Validate only.")

    class ImportImagesForm(forms.Form):
        root_path = forms.CharField(
            required=False,
            help_text="Server path containing folders named 'Title (Author)'."
        )
        archive_file = forms.FileField(
            required=False,
            help_text="Optional: upload a ZIP containing 'Title (Author)' folders.",
        )
        replace_existing = forms.BooleanField(
            required=False,
            initial=False,
            help_text="Replace existing images for matched books.",
        )
        dry_run = forms.BooleanField(required=False, initial=False, help_text="Validate only.")

        def clean(self):
            cleaned = super().clean()
            root_path = (cleaned.get("root_path") or "").strip()
            archive_file = cleaned.get("archive_file")
            if not root_path and not archive_file:
                raise forms.ValidationError("Provide either a server root path or a ZIP file.")
            if archive_file and not str(getattr(archive_file, "name", "")).lower().endswith(".zip"):
                raise forms.ValidationError("Uploaded file must be a .zip archive.")
            return cleaned

    def save_model(self, request, obj, form, change):
        original = None
        if change and obj.pk:
            original = Book.objects.filter(pk=obj.pk).first()

        super().save_model(request, obj, form, change)

        if not original:
            return

        if original.price != obj.price:
            _log_audit(
                request,
                "book_price_change",
                obj,
                {"price": {"from": str(original.price), "to": str(obj.price)}},
            )
        if original.stock != obj.stock:
            _log_audit(
                request,
                "book_stock_change",
                obj,
                {"stock": {"from": original.stock, "to": obj.stock}},
            )

    @admin.display(description="Discount")
    def discount_display(self, obj):
        if obj.discount_percentage > 0:
            return f"{obj.discount_percentage}% OFF"
        return "-"

    @admin.action(description="Bulk update price by percent")
    def bulk_update_price(self, request, queryset):
        raw = request.POST.get("percentage")
        if raw in (None, ""):
            self.message_user(request, "Enter a percentage value.", level="error")
            return
        try:
            pct = Decimal(str(raw))
        except Exception:
            self.message_user(request, "Invalid percentage.", level="error")
            return

        apply_to_mrp = bool(request.POST.get("apply_to_mrp"))
        factor = (Decimal("100") + pct) / Decimal("100")
        updated = 0

        for book in queryset:
            original_price = book.price
            book.price = (book.price * factor).quantize(Decimal("0.01"))
            if apply_to_mrp and book.mrp_price:
                book.mrp_price = (book.mrp_price * factor).quantize(Decimal("0.01"))
            book.save(update_fields=["price", "mrp_price"])
            if original_price != book.price:
                _log_audit(
                    request,
                    "book_price_change",
                    book,
                    {"price": {"from": str(original_price), "to": str(book.price)}},
                )
            updated += 1

        self.message_user(request, f"Updated {updated} books.")

    @admin.action(description="Mark as Featured")
    def mark_as_featured(self, request, queryset):
        updated = queryset.update(is_featured=True)
        self.message_user(request, f"Marked {updated} book(s) as featured.")

    @admin.action(description="Remove from Featured")
    def remove_from_featured(self, request, queryset):
        updated = queryset.update(is_featured=False)
        self.message_user(request, f"Removed {updated} book(s) from featured.")

    @admin.action(description="Mark as New Arrival")
    def mark_as_new_arrival(self, request, queryset):
        updated = queryset.update(is_new_arrival=True)
        self.message_user(request, f"Marked {updated} book(s) as new arrival.")

    @admin.action(description="Remove from New Arrival")
    def remove_from_new_arrival(self, request, queryset):
        updated = queryset.update(is_new_arrival=False)
        self.message_user(request, f"Removed {updated} book(s) from new arrival.")

    @admin.action(description="Mark as Bestseller")
    def mark_as_bestseller(self, request, queryset):
        updated = queryset.update(is_bestseller=True)
        self.message_user(request, f"Marked {updated} book(s) as bestseller.")

    @admin.action(description="Remove from Bestseller")
    def remove_from_bestseller(self, request, queryset):
        updated = queryset.update(is_bestseller=False)
        self.message_user(request, f"Removed {updated} book(s) from bestseller.")

    @admin.action(description="Mark as Trending")
    def mark_as_trending(self, request, queryset):
        updated = queryset.update(is_trending=True)
        self.message_user(request, f"Marked {updated} book(s) as trending.")

    @admin.action(description="Remove from Trending")
    def remove_from_trending(self, request, queryset):
        updated = queryset.update(is_trending=False)
        self.message_user(request, f"Removed {updated} book(s) from trending.")

    @admin.action(description="Activate books")
    def activate_books(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"Activated {updated} book(s).")

    @admin.action(description="Deactivate books")
    def deactivate_books(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"Deactivated {updated} book(s).")

    @admin.action(description="Export selected books to CSV")
    def export_books_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="books.csv"'
        writer = csv.writer(response)
        writer.writerow(["System ID", "Book Title", "Author", "Category", "Subject", "Price", "MRP", "Stock", "ISBN", "is_active"])
        for book in queryset:
            writer.writerow([
                book.system_id or "",
                book.title,
                book.author,
                book.category.name if book.category else "",
                " / ".join(book.subjects.values_list("name", flat=True)),
                book.price,
                book.mrp_price or "",
                book.stock,
                book.isbn or "",
                "true" if book.is_active else "false",
            ])
        return response

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export-books-xlsx/",
                self.admin_site.admin_view(self.download_books_export_xlsx_view),
                name="store_book_export_books_xlsx",
            ),
            path(
                "export-books/",
                self.admin_site.admin_view(self.download_books_export_csv_view),
                name="store_book_export_books",
            ),
            path(
                "import-books/sample-csv/",
                self.admin_site.admin_view(self.download_books_sample_csv_view),
                name="store_book_import_books_sample_csv",
            ),
            path(
                "import-books/",
                self.admin_site.admin_view(self.import_books_view),
                name="store_book_import_books",
            ),
            path(
                "import-images/",
                self.admin_site.admin_view(self.import_images_view),
                name="store_book_import_images",
            ),
        ]
        return custom_urls + urls

    def download_books_export_csv_view(self, request):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="books_export.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "System ID",
            "Book Title",
            "Author",
            "Category",
            "Subject",
            "Price",
            "MRP",
            "Stock",
            "ISBN No",
            "Description",
            "Published Year",
            "Binding",
            "Pages",
            "Weight",
            "is_active",
            "is_bestseller",
            "is_trending",
            "is_new_arrival",
            "is_featured",
        ])

        books = Book.objects.select_related("category").prefetch_related("subjects").order_by("title", "id")
        for book in books:
            writer.writerow([
                book.system_id or "",
                book.title,
                book.author,
                book.category.name if book.category else "",
                " / ".join(book.subjects.values_list("name", flat=True)),
                book.price,
                book.mrp_price or "",
                book.stock,
                book.isbn or "",
                book.description or "",
                book.published_year or "",
                book.binding or "",
                book.pages or "",
                book.weight or "",
                "true" if book.is_active else "false",
                "true" if book.is_bestseller else "false",
                "true" if book.is_trending else "false",
                "true" if book.is_new_arrival else "false",
                "true" if book.is_featured else "false",
            ])
        return response

    def download_books_export_xlsx_view(self, request):
        try:
            from openpyxl import Workbook
        except Exception:
            messages.error(request, "openpyxl is not installed on this server.")
            return redirect(reverse("admin:store_book_changelist"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Books"
        ws.append([
            "System ID",
            "Book Title",
            "Author",
            "Category",
            "Subject",
            "Price",
            "MRP",
            "Stock",
            "ISBN No",
            "Description",
            "Published Year",
            "Binding",
            "Pages",
            "Weight",
            "is_active",
            "is_bestseller",
            "is_trending",
            "is_new_arrival",
            "is_featured",
        ])

        books = Book.objects.select_related("category").prefetch_related("subjects").order_by("title", "id")
        for book in books:
            ws.append([
                book.system_id or "",
                book.title,
                book.author,
                book.category.name if book.category else "",
                " / ".join(book.subjects.values_list("name", flat=True)),
                str(book.price),
                str(book.mrp_price) if book.mrp_price is not None else "",
                book.stock,
                book.isbn or "",
                book.description or "",
                book.published_year or "",
                book.binding or "",
                book.pages or "",
                book.weight or "",
                "true" if book.is_active else "false",
                "true" if book.is_bestseller else "false",
                "true" if book.is_trending else "false",
                "true" if book.is_new_arrival else "false",
                "true" if book.is_featured else "false",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="books_export.xlsx"'
        wb.save(response)
        return response

    def download_books_sample_csv_view(self, request):
        columns = [
            "System ID",
            "Book Title",
            "Author",
            "Category",
            "Subject",
            "Price",
            "MRP",
            "Stock",
            "ISBN No",
            "Description",
            "Published Year",
            "Binding",
            "Pages",
            "Weight",
            "is_active",
            "is_bestseller",
            "is_trending",
            "is_new_arrival",
            "is_featured",
        ]
        sample_rows = [
            [
                "z7YrLMOPjbxmVdIM9U3g",
                "Makhzan-ul-Mufradat",
                "Hakim Example",
                "Unani Classics",
                "Ilmul Advia / Materia Medica",
                "450",
                "600",
                "25",
                "9780000000001",
                "Foundational Unani formulary reference.",
                "2024",
                "Paperback",
                "420",
                "0.45kg",
                "true",
                "true",
                "false",
                "true",
                "false",
            ]
        ]
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="books_import_sample.csv"'
        writer = csv.writer(response)
        writer.writerow(columns)
        for row in sample_rows:
            writer.writerow(row)
        return response

    def import_books_view(self, request):
        form = self.ImportBooksForm(request.POST or None, request.FILES or None)
        sample_csv_url = reverse("admin:store_book_import_books_sample_csv")
        if request.method == "POST" and form.is_valid():
            dry_run = form.cleaned_data["dry_run"]
            file = form.cleaned_data["file"]
            try:
                rows = _load_tabular_rows(file)
            except Exception as exc:
                messages.error(request, str(exc))
                return render(request, "admin/store/book/import_books.html", {
                    "form": form,
                    "sample_csv_url": sample_csv_url,
                })

            created = 0
            updated = 0
            errors = 0
            error_messages = []
            matched_by_system_id = 0
            matched_by_isbn = 0
            matched_by_title_author = 0

            with transaction.atomic():
                for row in rows:
                    try:
                        system_id = str(
                            row.get("system id")
                            or row.get("systemid")
                            or row.get("system id.")
                            or ""
                        ).strip()
                        title = str(row.get("book title") or row.get("title") or "").strip()
                        author = str(row.get("author") or "").strip()
                        if not title or not author:
                            errors += 1
                            error_messages.append(f"Row {row.get('_row_number', '?')}: title and author are required.")
                            continue

                        category_name = str(row.get("category") or row.get("caetgory") or "").strip()
                        subject_raw = row.get("subject") or row.get("subjects")
                        isbn = str(
                            row.get("isbn")
                            or row.get("isbn no")
                            or row.get("isbn no.")
                            or row.get("isbn number")
                            or ""
                        ).strip()
                        price_raw = row.get("price") or row.get("selling price")
                        mrp_raw = row.get("mrp") or row.get("mrp price") or row.get("mrp_price")
                        stock_raw = row.get("stock")
                        description = row.get("description")
                        published_year = row.get("published year") or row.get("published_year")
                        binding = row.get("binding")
                        pages = row.get("pages")
                        weight = row.get("weight")

                        book = None
                        is_created = False
                        match_reason = None
                        if system_id:
                            # When System ID is present, match only by System ID.
                            book = Book.objects.filter(system_id=system_id).first()
                            if book:
                                match_reason = "system_id"
                            else:
                                book = Book(
                                    title=title,
                                    author=author,
                                    system_id=system_id,
                                    isbn=isbn or "",
                                    price=Decimal("0.00"),
                                )
                                is_created = True
                        else:
                            if isbn:
                                book = Book.objects.filter(isbn=isbn).first()
                                if book:
                                    match_reason = "isbn"
                            if not book:
                                book, is_created = Book.objects.get_or_create(
                                    title=title,
                                    author=author,
                                    defaults={
                                        "system_id": None,
                                        "isbn": isbn or "",
                                        "price": Decimal("0.00"),
                                    },
                                )
                                if not is_created:
                                    match_reason = "title_author"

                        book.title = title
                        book.author = author

                        if category_name:
                            category, _ = Category.objects.get_or_create(name=category_name)
                            book.category = category

                        if system_id:
                            book.system_id = system_id

                        if isbn:
                            book.isbn = isbn

                        parsed_price = _coerce_decimal(price_raw)
                        if parsed_price is not None:
                            book.price = parsed_price

                        parsed_mrp = _coerce_decimal(mrp_raw)
                        if parsed_mrp is not None:
                            book.mrp_price = parsed_mrp

                        parsed_stock = _coerce_int(stock_raw)
                        if parsed_stock is not None:
                            book.stock = max(parsed_stock, 0)

                        if description not in (None, ""):
                            book.description = str(description)
                        if published_year not in (None, ""):
                            book.published_year = str(published_year).strip()
                        if binding not in (None, ""):
                            book.binding = str(binding).strip()
                        if pages not in (None, ""):
                            book.pages = str(pages).strip()
                        if weight not in (None, ""):
                            book.weight = str(weight).strip()

                        for flag_field in ("is_active", "is_bestseller", "is_trending", "is_new_arrival", "is_featured"):
                            parsed = _coerce_bool(row.get(flag_field))
                            if parsed is None:
                                parsed = _coerce_bool(row.get(flag_field.replace("_", " ")))
                            if parsed is not None:
                                setattr(book, flag_field, parsed)

                        if not dry_run:
                            book.save()

                        subjects = _split_multi(subject_raw)
                        if subjects and not dry_run:
                            subject_objs = []
                            for s in subjects:
                                obj, _ = Subject.objects.get_or_create(name=s)
                                subject_objs.append(obj)
                            book.subjects.set(subject_objs)

                        if is_created:
                            created += 1
                        else:
                            updated += 1
                            if match_reason == "system_id":
                                matched_by_system_id += 1
                            elif match_reason == "isbn":
                                matched_by_isbn += 1
                            elif match_reason == "title_author":
                                matched_by_title_author += 1
                    except Exception as exc:
                        errors += 1
                        error_messages.append(f"Row {row.get('_row_number', '?')}: {exc}")
                        continue

                if dry_run:
                    transaction.set_rollback(True)

            messages.success(
                request,
                "Import completed. "
                f"Created: {created}, Updated: {updated}, Errors: {errors}. "
                f"Updated by System ID: {matched_by_system_id}, "
                f"by ISBN: {matched_by_isbn}, "
                f"by Title+Author: {matched_by_title_author}"
                + (" (dry run)" if dry_run else ""),
            )
            for item in error_messages[:20]:
                messages.error(request, item)
            if len(error_messages) > 20:
                messages.error(request, f"Additional errors not shown: {len(error_messages) - 20}")

        return render(request, "admin/store/book/import_books.html", {
            "form": form,
            "sample_csv_url": sample_csv_url,
        })

    def _import_images_from_root(self, root, replace_existing=False, dry_run=False):
        from django.core.files import File

        created = 0
        skipped = 0
        errors = 0
        error_messages = []

        for folder in sorted(root.iterdir(), key=lambda p: p.name.lower()):
            if not folder.is_dir():
                continue
            match = re.match(r"^(.*)\((.*)\)$", folder.name)
            if not match:
                skipped += 1
                continue
            title = match.group(1).strip()
            author = match.group(2).strip()
            if not title or not author:
                skipped += 1
                continue

            book = Book.objects.filter(title=title, author=author).first()
            if not book:
                skipped += 1
                continue

            try:
                if replace_existing and not dry_run:
                    if book.main_cover:
                        book.main_cover.delete(save=False)
                    book.images.all().delete()

                image_files = sorted(
                    [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}],
                    key=lambda p: (not p.stem.isdigit(), int(p.stem) if p.stem.isdigit() else p.stem.lower()),
                )

                if not image_files:
                    skipped += 1
                    continue

                if not dry_run:
                    # 1.jpg as main cover
                    main = image_files[0]
                    with main.open("rb") as fh:
                        book.main_cover.save(main.name, File(fh), save=True)

                    for extra in image_files[1:]:
                        with extra.open("rb") as fh:
                            BookImage.objects.create(book=book, image=File(fh, name=extra.name))

                created += 1
            except Exception as exc:
                errors += 1
                error_messages.append(f"{folder.name}: {exc}")

        return created, skipped, errors, error_messages

    def _import_images_from_zip(self, archive_file, replace_existing=False, dry_run=False):
        from django.core.files.base import ContentFile

        created = 0
        skipped = 0
        errors = 0
        error_messages = []

        def image_sort_key(file_name):
            stem = file_name.rsplit(".", 1)[0]
            return (not stem.isdigit(), int(stem) if stem.isdigit() else stem.lower())

        grouped = {}
        try:
            archive_file.seek(0)
        except Exception:
            pass

        with zipfile.ZipFile(archive_file) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue

                path = str(info.filename or "").replace("\\", "/").strip("/")
                if not path:
                    continue

                parts = [part for part in path.split("/") if part]
                if len(parts) < 2:
                    continue

                file_name = parts[-1]
                suffix = f".{file_name.rsplit('.', 1)[-1].lower()}" if "." in file_name else ""
                if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
                    continue

                folder_name = None
                for segment in reversed(parts[:-1]):
                    if re.match(r"^(.*)\((.*)\)$", segment):
                        folder_name = segment
                        break
                if not folder_name:
                    continue

                grouped.setdefault(folder_name, []).append((file_name, info))

            for folder_name in sorted(grouped.keys(), key=lambda x: x.lower()):
                match = re.match(r"^(.*)\((.*)\)$", folder_name)
                if not match:
                    skipped += 1
                    continue

                title = match.group(1).strip()
                author = match.group(2).strip()
                if not title or not author:
                    skipped += 1
                    continue

                book = Book.objects.filter(title=title, author=author).first()
                if not book:
                    skipped += 1
                    continue

                image_items = sorted(grouped.get(folder_name, []), key=lambda item: image_sort_key(item[0]))
                if not image_items:
                    skipped += 1
                    continue

                try:
                    if replace_existing and not dry_run:
                        if book.main_cover:
                            book.main_cover.delete(save=False)
                        book.images.all().delete()

                    if not dry_run:
                        main_name, main_info = image_items[0]
                        main_data = zf.read(main_info)
                        book.main_cover.save(main_name, ContentFile(main_data, name=main_name), save=True)

                        for extra_name, extra_info in image_items[1:]:
                            extra_data = zf.read(extra_info)
                            BookImage.objects.create(
                                book=book,
                                image=ContentFile(extra_data, name=extra_name),
                            )

                    created += 1
                except Exception as exc:
                    errors += 1
                    error_messages.append(f"{folder_name}: {exc}")

        return created, skipped, errors, error_messages

    def import_images_view(self, request):
        form = self.ImportImagesForm(request.POST or None, request.FILES or None)
        if request.method == "POST" and form.is_valid():
            root_path = form.cleaned_data["root_path"].strip()
            archive_file = form.cleaned_data.get("archive_file")
            replace_existing = form.cleaned_data["replace_existing"]
            dry_run = form.cleaned_data["dry_run"]

            from pathlib import Path
            try:
                if archive_file:
                    try:
                        created, skipped, errors, error_messages = self._import_images_from_zip(
                            archive_file, replace_existing=replace_existing, dry_run=dry_run
                        )
                    except zipfile.BadZipFile:
                        messages.error(request, "Invalid ZIP file. Please upload a valid .zip archive.")
                        return render(request, "admin/store/book/import_images.html", {"form": form})
                else:
                    root = Path(root_path)
                    if not root.exists() or not root.is_dir():
                        messages.error(request, "Root path does not exist or is not a directory.")
                        return render(request, "admin/store/book/import_images.html", {"form": form})
                    created, skipped, errors, error_messages = self._import_images_from_root(
                        root, replace_existing=replace_existing, dry_run=dry_run
                    )
            except OSError as exc:
                messages.error(
                    request,
                    f"Import failed due to server storage/file limits: {exc}. "
                    "Try a smaller ZIP batch.",
                )
                return render(request, "admin/store/book/import_images.html", {"form": form})
            except Exception as exc:
                messages.error(request, f"Import failed: {exc}")
                return render(request, "admin/store/book/import_images.html", {"form": form})

            messages.success(
                request,
                f"Image import completed. Updated: {created}, Skipped: {skipped}, Errors: {errors}"
                + (" (dry run)" if dry_run else ""),
            )
            for item in error_messages[:20]:
                messages.error(request, item)
            if len(error_messages) > 20:
                messages.error(request, f"Additional errors not shown: {len(error_messages) - 20}")

        return render(request, "admin/store/book/import_images.html", {"form": form})
# ======================
# ORDER ITEM INLINE
# ======================

class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ("book", "price", "quantity", "allocated_quantity", "backordered_quantity")
    can_delete = False


# ======================
# ORDER ADMIN (GUEST-SAFE)
# ======================

@admin.register(Order)
class OrderAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "id",
        "full_name",
        "email_link",
        "total_cost_display",
        "sub_status",
        "coins_redeemed",
        "coins_earned_final",
        "coin_status",
        "packing_assignee",
        "shipping_assignee",
        "payment_method",
        "courier_service",
        "consignment_number",
        "tracking_link",
        "is_paid",
        "invoice_link",
        "city",
        "status",
        "created_at",
    )

    list_filter = ("status", "is_paid", "created_at", "city", "packing_assignee", "shipping_assignee", "courier_service")
    search_fields = ("full_name", "email", "city", "user__username", "consignment_number")
    list_select_related = ("user", "coupon", "packing_assignee", "shipping_assignee")
    autocomplete_fields = ("coupon",)
    readonly_fields = (
        "user",
        "created_at",
        "subtotal",
        "discount_amount",
        "gst_rate",
        "gst_amount",
        "shipping_amount",
        "total_cost",
        "payment_method",
        "razorpay_order_id",
        "razorpay_payment_id",
        "razorpay_signature",
        "coins_redeemed",
        "coins_earned_estimate",
        "coins_earned_final",
        "coin_status",
    )

    date_hierarchy = "created_at"
    inlines = [OrderItemInline]
    ordering = ("-created_at",)
    actions = (
        "set_status_processing",
        "set_status_packed",
        "set_status_shipped",
        "export_orders_csv",
        "force_credit_coins",
        "cancel_pending_coin_credit",
    )

    def save_model(self, request, obj, form, change):
        original = None
        if change and obj.pk:
            original = Order.objects.filter(pk=obj.pk).first()

        super().save_model(request, obj, form, change)

        if not original:
            return

        if original.is_paid != obj.is_paid:
            _log_audit(
                request,
                "order_paid_change",
                obj,
                {"is_paid": {"from": original.is_paid, "to": obj.is_paid}},
            )
        if original.status != obj.status:
            _log_audit(
                request,
                "order_status_change",
                obj,
                {"status": {"from": original.status, "to": obj.status}},
            )
            if obj.user_id and obj.status == "Delivered":
                queue_order_pending_rewards(obj)
                process_due_pending_rewards_for_user(obj.user)
            if obj.user_id and obj.status == "Cancelled":
                txs = IKSWalletTransaction.objects.filter(order=obj, status="pending")
                if txs.exists():
                    wallet = IKSWallet.objects.filter(user=obj.user).first()
                    if wallet:
                        pending_sum = txs.aggregate(total=Sum("coins"))["total"] or 0
                        wallet.pending_balance = max(wallet.pending_balance - max(int(pending_sum), 0), 0)
                        wallet.save(update_fields=["pending_balance", "updated_at"])
                    txs.update(status="cancelled", completed_at=timezone.now(), note="Auto-cancelled with order")
                    obj.coin_status = "cancelled"
                    obj.save(update_fields=["coin_status"])

    @admin.display(description="Email")
    def email_link(self, obj):
        return format_html(
            '<a href="mailto:{}">{}</a>',
            obj.email,
            obj.email
        )

    @admin.display(description="Total Cost")
    def total_cost_display(self, obj):
        return f"₹{obj.total_cost}"

    @admin.display(description="Invoice")
    def invoice_link(self, obj):
        return format_html(
            '<a href="/invoice/{}/" target="_blank">View</a>',
            obj.id,
        )

    @admin.display(description="Tracking")
    def tracking_link(self, obj):
        if obj.courier_service == "india_post" and obj.consignment_number:
            return format_html(
                '<a href="{}" target="_blank" rel="noopener noreferrer">India Post</a>',
                obj.tracking_url,
            )
        return "—"



    @admin.action(description="Export selected orders to CSV")
    def export_orders_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="orders.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "ID",
            "Name",
            "Email",
            "City",
            "Total",
            "Paid",
            "Created",
        ])
        for order in queryset:
            writer.writerow([
                order.id,
                order.full_name,
                order.email,
                order.city,
                order.total_cost,
                "Yes" if order.is_paid else "No",
                order.created_at,
            ])
        return response

    def _bulk_update_status(self, request, queryset, target_status):
        updated = 0
        for order in queryset:
            if order.status == target_status:
                continue
            old_status = order.status
            order.status = target_status
            order.save(update_fields=["status"])
            _log_audit(
                request,
                "order_status_change",
                order,
                {"status": {"from": old_status, "to": target_status}},
            )
            updated += 1
        self.message_user(request, f"Updated {updated} order(s) to {target_status}.")

    @admin.action(description="Set status: Processing")
    def set_status_processing(self, request, queryset):
        self._bulk_update_status(request, queryset, "Processing")

    @admin.action(description="Set status: Packed")
    def set_status_packed(self, request, queryset):
        self._bulk_update_status(request, queryset, "Packed")

    @admin.action(description="Set status: Shipped")
    def set_status_shipped(self, request, queryset):
        self._bulk_update_status(request, queryset, "Shipped")

    @admin.action(description="Force credit pending coins now")
    def force_credit_coins(self, request, queryset):
        updated = 0
        for order in queryset.select_related("user"):
            if not order.user_id:
                continue
            if order.coin_status == "pending":
                order.coin_release_date = timezone.now()
                order.coin_manual_override = True
                order.save(update_fields=["coin_release_date", "coin_manual_override"])
                process_due_pending_rewards_for_user(order.user)
                updated += 1
        self.message_user(request, f"Forced coin release for {updated} order(s).")

    @admin.action(description="Cancel pending coin credits")
    def cancel_pending_coin_credit(self, request, queryset):
        cancelled = 0
        for order in queryset.select_related("user"):
            if order.coin_status != "pending":
                continue
            txs = IKSWalletTransaction.objects.filter(order=order, status="pending")
            if not txs.exists():
                continue
            if order.user_id:
                wallet = IKSWallet.objects.filter(user=order.user).first()
                if wallet:
                    pending_sum = txs.aggregate(total=Sum("coins"))["total"] or 0
                    wallet.pending_balance = max(wallet.pending_balance - max(int(pending_sum), 0), 0)
                    wallet.save(update_fields=["pending_balance", "updated_at"])
            txs.update(status="cancelled", completed_at=timezone.now(), note="Cancelled by admin")
            order.coin_status = "cancelled"
            order.coin_manual_override = True
            order.save(update_fields=["coin_status", "coin_manual_override"])
            cancelled += 1
        self.message_user(request, f"Cancelled pending coin credit for {cancelled} order(s).")
# ======================
# CART ADMIN
# ======================

class CartItemInline(admin.TabularInline):
    model = CartItem
    extra = 0


class CartBundleItemInline(admin.TabularInline):
    model = CartBundleItem
    extra = 0


@admin.register(Cart)
class CartAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("user", "created_at")
    search_fields = ("user__username",)
    inlines = [CartItemInline, CartBundleItemInline]


# ======================
# WISHLIST ADMIN
# ======================

@admin.register(Wishlist)
class WishlistAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("user", "book", "created_at")
    search_fields = ("user__username", "book__title")
    list_select_related = ("user", "book")


# ======================
# REVIEW ADMIN
# ======================

@admin.register(Review)
class ReviewAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "book",
        "user",
        "rating",
        "short_comment",
        "created_at",
    )

    list_filter = ("rating", "created_at")
    search_fields = ("book__title", "user__username", "comment")
    list_select_related = ("book", "user")
    ordering = ("-created_at",)
    readonly_fields = ("created_at",)

    @admin.display(description="Comment")
    def short_comment(self, obj):
        return (obj.comment[:50] + "…") if obj.comment else "—"


# ======================
# COUPON ADMIN
# ======================

@admin.register(Coupon)
class CouponAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "code",
        "discount_type",
        "value",
        "minimum_order_amount",
        "active",
        "expiry_date",
        "is_valid_coupon",
        "created_at",
    )

    list_filter = ("discount_type", "active")
    search_fields = ("code",)
    ordering = ("-created_at",)

    fieldsets = (
        ("Coupon Details", {
            "fields": ("code", "discount_type", "value", "minimum_order_amount", "valid_categories")
        }),
        ("Status & Expiry", {
            "fields": ("expiry_date", "active")
        }),
    )

    readonly_fields = ("created_at",)

    @admin.display(boolean=True, description="Valid")
    def is_valid_coupon(self, obj):
        if not obj.active:
            return False
        if obj.expiry_date and obj.expiry_date < timezone.now().date():
            return False
        return True

    def save_model(self, request, obj, form, change):
        if obj.expiry_date and obj.expiry_date < timezone.now().date():
            obj.active = False
        super().save_model(request, obj, form, change)


class UnaniTermActionForm(ActionForm):
    reference_source = forms.ModelChoiceField(
        queryset=UnaniReferenceSource.objects.none(),
        required=False,
        label="Reference (for assign action)",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["reference_source"].queryset = UnaniReferenceSource.objects.filter(
            is_active=True
        ).order_by("name")


class UnaniTermAdminForm(forms.ModelForm):
    class Meta:
        model = UnaniTerm
        fields = "__all__"

    def clean_arabic_script(self):
        return normalize_script_text(self.cleaned_data.get("arabic_script"))

    def clean_transliteration(self):
        value = normalize_transliteration_text(self.cleaned_data.get("transliteration"))
        validate_transliteration_style(value)
        return value


class UnaniSectionListFilter(admin.SimpleListFilter):
    title = "section"
    parameter_name = "section"
    uncategorized_value = "__blank__"

    def lookups(self, request, model_admin):
        base_qs = model_admin.get_queryset(request).annotate(section_clean=Trim("section"))
        sections = (
            base_qs.exclude(section_clean="")
            .values_list("section_clean", flat=True)
            .distinct()
            .order_by("section_clean")
        )
        lookups = [(section, section) for section in sections]
        if base_qs.filter(section_clean="").exists():
            lookups.append((self.uncategorized_value, "(Uncategorized)"))
        return lookups

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset
        filtered = queryset.annotate(section_clean=Trim("section"))
        if value == self.uncategorized_value:
            return filtered.filter(section_clean="")
        return filtered.filter(section_clean=value)


@admin.register(UnaniTerm)
class UnaniTermAdmin(BulkImportAdminMixin, ProductivityAdminMixin, admin.ModelAdmin):
    form = UnaniTermAdminForm
    action_form = UnaniTermActionForm
    change_list_template = "admin/store/unaniterm/change_list.html"
    list_display = ("english_term", "reference_source", "section", "is_published", "updated_at")
    search_fields = (
        "arabic_script",
        "arabic_script_normalized",
        "transliteration",
        "transliteration_normalized",
        "english_term",
        "english_term_normalized",
        "description",
    )
    list_filter = (UnaniSectionListFilter, "reference_source", "is_published")
    prepopulated_fields = {"slug": ("english_term",)}
    ordering = ("english_term",)
    readonly_fields = ("created_at", "updated_at")
    list_per_page = 50
    actions = ("publish_terms", "unpublish_terms", "assign_reference_source", "clear_reference_source")
    bulk_import_title = "Bulk Import Unani Terms"
    bulk_import_help = (
        "Required: english_term. Optional: description, transliteration, arabic_script, "
        "section, reference_source, slug, is_published."
    )
    bulk_import_overflow_merge_column = "description"
    bulk_import_columns = (
        "english_term",
        "description",
        "transliteration",
        "arabic_script",
        "section",
        "reference_source",
        "slug",
        "is_published",
    )
    bulk_import_sample_rows = (
        {
            "english_term": "Tamamiyya Asbab",
            "description": "Causes related to functions.",
            "transliteration": "Tamamiyya Asbab",
            "arabic_script": "\u0627\u0633\u0628\u0627\u0628 \u062a\u0627\u0645\u06cc\u06c1",
            "section": "General Terms",
            "reference_source": "Qarabadin-e-Qadri",
            "slug": "tamamiyya-asbab",
            "is_published": "true",
        },
    )

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        if search_term:
            normalized_script = normalize_script_text(search_term)
            normalized_text = normalize_latin_search_text(search_term)
            queryset = queryset.annotate(
                search_priority=Case(
                    When(arabic_script_normalized=normalized_script, then=Value(1)),
                    When(arabic_script_normalized__startswith=normalized_script, then=Value(2)),
                    When(transliteration_normalized__icontains=normalized_text, then=Value(3)),
                    When(english_term_normalized__icontains=normalized_text, then=Value(4)),
                    When(description__icontains=search_term, then=Value(5)),
                    default=Value(6),
                    output_field=IntegerField(),
                )
            ).order_by("search_priority", "english_term")
        return queryset, use_distinct

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        since = timezone.now() - timedelta(days=7)
        top_searches = list(
            DictionaryQueryLog.objects.filter(created_at__gte=since)
            .exclude(query="")
            .values("query")
            .annotate(total=Count("id"))
            .order_by("-total", "query")[:10]
        )
        zero_result_searches = list(
            DictionaryQueryLog.objects.filter(created_at__gte=since, results_count=0)
            .exclude(query="")
            .values("query")
            .annotate(total=Count("id"))
            .order_by("-total", "query")[:10]
        )
        most_opened_terms = list(
            DictionaryTermOpenLog.objects.filter(created_at__gte=since)
            .values("term_id", "term__english_term")
            .annotate(total=Count("id"))
            .order_by("-total", "term__english_term")[:10]
        )
        extra_context["dictionary_analytics"] = {
            "since": since,
            "top_searches": top_searches,
            "zero_result_searches": zero_result_searches,
            "most_opened_terms": most_opened_terms,
        }
        return super().changelist_view(request, extra_context=extra_context)

    @admin.action(description="Publish selected terms")
    def publish_terms(self, request, queryset):
        count = queryset.update(is_published=True)
        self.message_user(request, f"{count} term(s) published.")

    @admin.action(description="Unpublish selected terms")
    def unpublish_terms(self, request, queryset):
        count = queryset.update(is_published=False)
        self.message_user(request, f"{count} term(s) unpublished.")

    @admin.action(description="Assign selected reference source")
    def assign_reference_source(self, request, queryset):
        reference_id = request.POST.get("reference_source")
        if not reference_id:
            self.message_user(
                request,
                "Choose a reference from the dropdown before running this action.",
                level=messages.ERROR,
            )
            return
        reference = UnaniReferenceSource.objects.filter(pk=reference_id, is_active=True).first()
        if not reference:
            self.message_user(
                request,
                "Selected reference is invalid or inactive.",
                level=messages.ERROR,
            )
            return
        count = queryset.update(reference_source=reference)
        self.message_user(request, f"{count} term(s) updated with reference '{reference.name}'.")

    @admin.action(description="Clear reference source")
    def clear_reference_source(self, request, queryset):
        count = queryset.update(reference_source=None)
        self.message_user(request, f"Reference source cleared for {count} term(s).")

    def import_row(self, row_data):
        english_term = str(row_data.get("english term") or row_data.get("english_term") or "").strip()
        if not english_term:
            raise ValueError("english_term is required.")
        if len(english_term) > 255:
            raise ValueError(f"english_term exceeds 255 characters (got {len(english_term)}).")

        term, created = UnaniTerm.objects.get_or_create(
            english_term=english_term,
            defaults={"description": str(row_data.get("description") or "").strip()},
        )

        for field_name, headers in (
            ("description", ("description",)),
            ("transliteration", ("transliteration",)),
            ("arabic_script", ("arabic script", "arabic_script")),
            ("section", ("section",)),
            ("slug", ("slug",)),
        ):
            value = None
            for header in headers:
                if row_data.get(header) not in (None, ""):
                    value = str(row_data.get(header)).strip()
                    break
            if value is not None:
                setattr(term, field_name, value)

        reference_name = str(
            row_data.get("reference source")
            or row_data.get("reference_source")
            or row_data.get("reference")
            or ""
        ).strip()
        if reference_name:
            reference, _ = UnaniReferenceSource.objects.get_or_create(name=reference_name)
            term.reference_source = reference

        term.arabic_script = normalize_script_text(term.arabic_script)
        term.transliteration = normalize_transliteration_text(term.transliteration)
        bad_chars = transliteration_invalid_chars(term.transliteration)
        if bad_chars:
            bad_display = ", ".join(repr(ch) for ch in bad_chars[:8])
            raise ValueError(
                "Transliteration has unsupported characters. "
                f"Use one style (Latin + diacritics). Found: {bad_display}"
            )

        char_limits = {
            "transliteration": 255,
            "arabic_script": 255,
            "section": 120,
            "slug": 255,
        }
        for field_name, max_len in char_limits.items():
            value = getattr(term, field_name, "") or ""
            if len(value) > max_len:
                raise ValueError(f"{field_name} exceeds {max_len} characters (got {len(value)}).")

        parsed_publish = _coerce_bool(row_data.get("is published") if "is published" in row_data else row_data.get("is_published"))
        if parsed_publish is not None:
            term.is_published = parsed_publish

        term.save()
        return "created" if created else "updated"


@admin.register(ClassicalWeightUnit)
class ClassicalWeightUnitAdmin(BulkImportAdminMixin, ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("classical_weight", "metric_weight", "grams_value", "display_order", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("classical_weight", "metric_weight", "source_note")
    list_editable = ("display_order", "is_active")
    ordering = ("display_order", "id")
    readonly_fields = ("created_at", "updated_at")
    change_list_template = BulkImportAdminMixin.bulk_import_changelist_template
    bulk_import_title = "Bulk Import Classical Weight Units"
    bulk_import_help = "Required: classical_weight, metric_weight, grams_value. Optional: display_order, is_active, source_note."
    bulk_import_columns = ("classical_weight", "metric_weight", "grams_value", "display_order", "is_active", "source_note")
    bulk_import_sample_rows = (
        {
            "classical_weight": "Ratti",
            "metric_weight": "121.5mg",
            "grams_value": "0.1215",
            "display_order": "1",
            "is_active": "true",
            "source_note": "Traditional Unani reference",
        },
    )
    fieldsets = (
        ("Unit", {
            "fields": ("classical_weight", "metric_weight", "grams_value")
        }),
        ("Display", {
            "fields": ("display_order", "is_active")
        }),
        ("Reference", {
            "fields": ("source_note",)
        }),
        ("Timestamps", {
            "fields": ("created_at", "updated_at")
        }),
    )

    def import_row(self, row_data):
        classical_weight = str(row_data.get("classical weight") or row_data.get("classical_weight") or "").strip()
        metric_weight = str(row_data.get("metric weight") or row_data.get("metric_weight") or "").strip()
        grams_value = _coerce_decimal(row_data.get("grams value") if "grams value" in row_data else row_data.get("grams_value"))
        if not classical_weight or not metric_weight or grams_value is None:
            raise ValueError("classical_weight, metric_weight, and grams_value are required.")

        unit, created = ClassicalWeightUnit.objects.get_or_create(
            classical_weight=classical_weight,
            defaults={
                "metric_weight": metric_weight,
                "grams_value": grams_value,
            },
        )

        unit.metric_weight = metric_weight
        unit.grams_value = grams_value

        parsed_order = _coerce_int(row_data.get("display order") if "display order" in row_data else row_data.get("display_order"))
        if parsed_order is not None:
            unit.display_order = max(parsed_order, 0)

        parsed_active = _coerce_bool(row_data.get("is active") if "is active" in row_data else row_data.get("is_active"))
        if parsed_active is not None:
            unit.is_active = parsed_active

        source_note = row_data.get("source note") if "source note" in row_data else row_data.get("source_note")
        if source_note not in (None, ""):
            unit.source_note = str(source_note).strip()

        unit.save()
        return "created" if created else "updated"


# ======================
# SYLLABUS PDF ADMIN
# ======================

@admin.register(SyllabusPDF)
class SyllabusPDFAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("title", "category", "semester", "uploaded_at")
    list_filter = ("category", "semester")
    search_fields = ("title",)
    ordering = ("category", "semester")

    fieldsets = (
        ("File Information", {
            "fields": ("title", "pdf_file")
        }),
        ("Classification", {
            "fields": ("category", "semester")
        }),
    )


# ======================
# GALLERY ADMIN
# ======================

@admin.register(GalleryItem)
class GalleryItemAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("event", "media_type", "title", "order", "created_at")
    list_filter = ("event", "media_type")
    search_fields = ("event", "title", "media")
    ordering = ("event", "order", "created_at")
    readonly_fields = ("created_at",)


# ======================
# USER PROFILE ADMIN
# ======================

@admin.register(UserProfile)
class UserProfileAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("user", "phone", "city", "company", "iks_follow_instagram", "iks_follow_facebook", "updated_at")
    search_fields = ("user__username", "user__email", "phone", "city", "company")
    ordering = ("-updated_at",)


@admin.register(UserAddress)
class UserAddressAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("user", "label", "city", "zip_code", "is_default", "created_at")
    list_filter = ("is_default", "city")
    search_fields = ("user__username", "user__email", "label", "address", "city", "zip_code")
    ordering = ("-created_at",)


@admin.register(SearchQueryLog)
class SearchQueryLogAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("query", "results_count", "user", "created_at")
    list_filter = ("results_count", "created_at")
    search_fields = ("query", "category_slug", "subject_slug", "ip_address")
    ordering = ("-created_at",)
    date_hierarchy = "created_at"


@admin.register(UnaniReferenceSource)
class UnaniReferenceSourceAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("name", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("name", "citation", "source_url")
    ordering = ("name",)


@admin.register(DictionaryQueryLog)
class DictionaryQueryLogAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("query", "results_count", "section", "letter", "user", "created_at")
    list_filter = ("results_count", "section", "letter", "created_at")
    search_fields = ("query", "normalized_query", "ip_address")
    ordering = ("-created_at",)
    date_hierarchy = "created_at"
    readonly_fields = ("query", "normalized_query", "results_count", "section", "letter", "user", "ip_address", "created_at")


@admin.register(DictionaryTermOpenLog)
class DictionaryTermOpenLogAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("term", "user", "ip_address", "created_at")
    list_filter = ("created_at",)
    search_fields = ("term__english_term", "term__arabic_script", "term__transliteration", "ip_address")
    ordering = ("-created_at",)
    date_hierarchy = "created_at"
    readonly_fields = ("term", "user", "ip_address", "created_at")


@admin.register(AuditLog)
class AuditLogAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("action", "model_name", "object_id", "user", "created_at")
    list_filter = ("action", "model_name", "created_at")
    search_fields = ("object_repr", "user__username", "user__email")
    readonly_fields = ("user", "action", "model_name", "object_id", "object_repr", "changes", "created_at")
    ordering = ("-created_at",)
    date_hierarchy = "created_at"


@admin.register(SiteSettings)
class SiteSettingsAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("id", "sales_offers_label", "sales_offers_enabled", "is_active", "created_at")
    list_filter = ("is_active", "created_at")
    ordering = ("-created_at",)
    fieldsets = (
        ("Branding", {
            "fields": ("background_image", "loader_logo")
        }),
        ("Navbar", {
            "fields": ("sales_offers_label", "sales_offers_enabled")
        }),
        ("Status", {
            "fields": ("is_active",)
        }),
    )


@admin.register(IKSCoinsSettings)
class IKSCoinsSettingsAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "name",
        "is_active",
        "program_enabled",
        "earn_percentage",
        "max_coins_per_order",
        "monthly_earning_cap",
        "redemption_percentage_limit",
        "credit_delay_days",
        "updated_at",
    )
    list_filter = ("is_active", "program_enabled")
    ordering = ("-updated_at",)


@admin.register(IKSWallet)
class IKSWalletAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = (
        "user",
        "balance",
        "pending_balance",
        "total_earned",
        "total_redeemed",
        "monthly_earned",
        "is_frozen",
        "is_earning_blocked",
        "updated_at",
    )
    search_fields = ("user__username", "user__email")
    list_filter = ("is_frozen", "is_earning_blocked")
    readonly_fields = ("updated_at",)
    actions = ("freeze_wallets", "unfreeze_wallets", "block_earning", "unblock_earning", "reset_wallets")

    @admin.action(description="Freeze selected wallets")
    def freeze_wallets(self, request, queryset):
        updated = queryset.update(is_frozen=True)
        self.message_user(request, f"Frozen {updated} wallet(s).")

    @admin.action(description="Unfreeze selected wallets")
    def unfreeze_wallets(self, request, queryset):
        updated = queryset.update(is_frozen=False)
        self.message_user(request, f"Unfroze {updated} wallet(s).")

    @admin.action(description="Block selected users from earning")
    def block_earning(self, request, queryset):
        updated = queryset.update(is_earning_blocked=True)
        self.message_user(request, f"Blocked earning for {updated} wallet(s).")

    @admin.action(description="Unblock selected users from earning")
    def unblock_earning(self, request, queryset):
        updated = queryset.update(is_earning_blocked=False)
        self.message_user(request, f"Unblocked earning for {updated} wallet(s).")

    @admin.action(description="Reset selected wallets to zero")
    def reset_wallets(self, request, queryset):
        count = 0
        for wallet in queryset:
            delta = -int(wallet.balance)
            if delta:
                manual_adjust_wallet(wallet, delta, note="Admin wallet reset")
            wallet.pending_balance = 0
            wallet.monthly_earned = 0
            wallet.save(update_fields=["pending_balance", "monthly_earned", "updated_at"])
            count += 1
        self.message_user(request, f"Reset {count} wallet(s).")


@admin.register(IKSWalletTransaction)
class IKSWalletTransactionAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("id", "wallet", "tx_type", "status", "coins", "order", "book", "release_date", "created_at")
    list_filter = ("tx_type", "status", "created_at")
    search_fields = ("wallet__user__username", "wallet__user__email", "note")
    readonly_fields = ("created_at", "completed_at")
    autocomplete_fields = ("wallet", "order", "book")

    def save_model(self, request, obj, form, change):
        is_new = not change
        super().save_model(request, obj, form, change)
        if not is_new:
            return
        if obj.tx_type != "manual_adjustment" or obj.status != "completed" or not obj.wallet_id:
            return
        wallet = obj.wallet
        if obj.coins >= 0:
            wallet.balance += obj.coins
            wallet.total_earned += obj.coins
        else:
            wallet.balance = max(wallet.balance + obj.coins, 0)
            wallet.total_redeemed += abs(obj.coins)
        wallet.save(update_fields=["balance", "total_earned", "total_redeemed", "updated_at"])
        if not obj.completed_at:
            obj.completed_at = timezone.now()
            obj.save(update_fields=["completed_at"])

@admin.register(Subject)
class SubjectAdmin(BulkImportAdminMixin, ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("name", "slug", "is_active")
    search_fields = ("name",)
    list_filter = ("is_active",)
    prepopulated_fields = {"slug": ("name",)}
    change_list_template = BulkImportAdminMixin.bulk_import_changelist_template
    bulk_import_title = "Bulk Import Subjects"
    bulk_import_help = "Required: name. Optional: slug, is_active."
    bulk_import_columns = ("name", "slug", "is_active")
    bulk_import_sample_rows = (
        {"name": "Ilmul Advia", "slug": "ilmul-advia", "is_active": "true"},
        {"name": "Tashreeh", "slug": "tashreeh", "is_active": "true"},
    )

    def import_row(self, row_data):
        name = str(row_data.get("name") or "").strip()
        if not name:
            raise ValueError("name is required.")

        subject, created = Subject.objects.get_or_create(name=name)
        slug = str(row_data.get("slug") or "").strip()
        if slug:
            subject.slug = slug
        parsed_active = _coerce_bool(row_data.get("is active") if "is active" in row_data else row_data.get("is_active"))
        if parsed_active is not None:
            subject.is_active = parsed_active
        subject.save()
        return "created" if created else "updated"


@admin.register(PublishWithUsSubmission)
class PublishWithUsSubmissionAdmin(ProductivityAdminMixin, admin.ModelAdmin):
    list_display = ("title", "author_name", "email", "phone", "created_at")
    search_fields = ("title", "author_name", "email", "phone")
    readonly_fields = ("created_at",)
    ordering = ("-created_at",)


class BannerAdminForm(forms.ModelForm):
    class Meta:
        model = Banner
        fields = "__all__"
        widgets = {
            "desktop_crop_x": forms.NumberInput(attrs={"min": 0}),
            "desktop_crop_y": forms.NumberInput(attrs={"min": 0}),
            "desktop_crop_width": forms.NumberInput(attrs={"min": 0}),
            "desktop_crop_height": forms.NumberInput(attrs={"min": 0}),
            "mobile_crop_x": forms.NumberInput(attrs={"min": 0}),
            "mobile_crop_y": forms.NumberInput(attrs={"min": 0}),
            "mobile_crop_width": forms.NumberInput(attrs={"min": 0}),
            "mobile_crop_height": forms.NumberInput(attrs={"min": 0}),
        }


@admin.register(Banner)
class BannerAdmin(BulkImportAdminMixin, ProductivityAdminMixin, admin.ModelAdmin):
    form = BannerAdminForm
    list_display = (
        "preview",
        "title",
        "category",
        "order",
        "is_active",
        "show_on_mobile",
        "show_on_desktop",
        "focal_x",
        "focal_y",
        "desktop_crop_width",
        "desktop_crop_height",
        "mobile_crop_width",
        "mobile_crop_height",
        "mobile_height",
        "tablet_height",
    )
    list_editable = (
        "order",
        "is_active",
        "show_on_mobile",
        "show_on_desktop",
        "focal_x",
        "focal_y",
        "desktop_crop_width",
        "desktop_crop_height",
        "mobile_crop_width",
        "mobile_crop_height",
        "mobile_height",
        "tablet_height",
    )
    search_fields = ("title", "headline", "subheadline")
    ordering = ("order", "id")
    change_list_template = BulkImportAdminMixin.bulk_import_changelist_template
    bulk_import_title = "Bulk Import Banners"
    bulk_import_help = (
        "Required for create: image. Recommended identifiers: id or title+order. "
        "Category fields accept category names."
    )
    bulk_import_columns = (
        "id",
        "title",
        "headline",
        "subheadline",
        "image",
        "category",
        "cta_text",
        "cta_category",
        "order",
        "is_active",
        "show_on_mobile",
        "show_on_desktop",
        "focal_x",
        "focal_y",
        "desktop_crop_x",
        "desktop_crop_y",
        "desktop_crop_width",
        "desktop_crop_height",
        "mobile_crop_x",
        "mobile_crop_y",
        "mobile_crop_width",
        "mobile_crop_height",
        "mobile_height",
        "tablet_height",
    )
    bulk_import_sample_rows = (
        {
            "id": "",
            "title": "Homepage Hero",
            "headline": "Welcome to Idara",
            "subheadline": "Discover curated Unani resources",
            "image": "banners/sample-banner.jpg",
            "category": "Unani Classics",
            "cta_text": "Shop Now",
            "cta_category": "Unani Classics",
            "order": "1",
            "is_active": "true",
            "show_on_mobile": "true",
            "show_on_desktop": "true",
            "focal_x": "50",
            "focal_y": "50",
            "desktop_crop_x": "0",
            "desktop_crop_y": "0",
            "desktop_crop_width": "0",
            "desktop_crop_height": "0",
            "mobile_crop_x": "0",
            "mobile_crop_y": "0",
            "mobile_crop_width": "0",
            "mobile_crop_height": "0",
            "mobile_height": "360",
            "tablet_height": "420",
        },
    )

    readonly_fields = ("crop_box_tool",)

    fieldsets = (
        ("Banner", {
            "fields": ("title", "headline", "subheadline", "image", "category", "order", "is_active", "show_on_mobile", "show_on_desktop")
        }),
        ("CTA", {
            "fields": ("cta_text", "cta_category")
        }),
        ("Crop Box", {
            "fields": ("crop_box_tool",)
        }),
        ("Desktop Crop Box", {
            "fields": ("desktop_crop_x", "desktop_crop_y", "desktop_crop_width", "desktop_crop_height")
        }),
        ("Mobile Crop Box", {
            "fields": ("mobile_crop_x", "mobile_crop_y", "mobile_crop_width", "mobile_crop_height")
        }),
        ("Display Position & Height", {
            "fields": ("focal_x", "focal_y", "mobile_height", "tablet_height")
        }),
    )

    class Media:
        css = {"all": ("admin/css/banner_crop_box.css",)}
        js = ("admin/js/banner_crop_box.js",)

    @admin.display(description="Preview")
    def preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:48px;width:86px;object-fit:cover;border-radius:6px;border:1px solid #243149;" />',
                obj.image.url,
            )
        return "?"

    @admin.display(description="Crop Tool")
    def crop_box_tool(self, obj):
        if not obj or not obj.image:
            return "Save with an image first, then use crop box."
        return format_html(
            """
            <div class="banner-crop-tool-wrap">
              <p class="help">
                Select Desktop or Mobile mode, then drag on image to set crop box.
                Values are saved in pixel coordinates.
              </p>
              <div class="banner-crop-mode">
                <button type="button" class="button" data-crop-mode="desktop">Desktop</button>
                <button type="button" class="button" data-crop-mode="mobile">Mobile</button>
              </div>
              <div class="banner-crop-stage" id="banner-crop-stage">
                <img id="banner-crop-image" src="{}" alt="Banner crop source" />
                <div id="banner-crop-rect"></div>
              </div>
            </div>
            """,
            obj.image.url,
        )

    def import_row(self, row_data):
        banner_id = _coerce_int(row_data.get("id"))
        title = str(row_data.get("title") or "").strip()
        order_val = _coerce_int(row_data.get("order"), default=0)

        banner = None
        if banner_id:
            banner = Banner.objects.filter(pk=banner_id).first()
        if banner is None and title:
            banner = Banner.objects.filter(title=title, order=order_val).first()

        created = banner is None
        if banner is None:
            banner = Banner(order=order_val)

        for field_name in ("title", "headline", "subheadline", "cta_text"):
            value = row_data.get(field_name)
            if value not in (None, ""):
                setattr(banner, field_name, str(value).strip())

        image_val = row_data.get("image")
        if image_val not in (None, ""):
            banner.image = str(image_val).strip()
        elif created and not banner.image:
            raise ValueError("image is required for new banner rows.")

        category_name = str(row_data.get("category") or "").strip()
        if category_name:
            banner.category, _ = Category.objects.get_or_create(name=category_name)

        cta_category_name = str(row_data.get("cta category") or row_data.get("cta_category") or "").strip()
        if cta_category_name:
            banner.cta_category, _ = Category.objects.get_or_create(name=cta_category_name)

        int_fields = (
            "order",
            "focal_x",
            "focal_y",
            "desktop_crop_x",
            "desktop_crop_y",
            "desktop_crop_width",
            "desktop_crop_height",
            "mobile_crop_x",
            "mobile_crop_y",
            "mobile_crop_width",
            "mobile_crop_height",
            "mobile_height",
            "tablet_height",
        )
        for field_name in int_fields:
            parsed = _coerce_int(row_data.get(field_name))
            if parsed is not None:
                setattr(banner, field_name, parsed)

        bool_map = {
            "is_active": row_data.get("is active") if "is active" in row_data else row_data.get("is_active"),
            "show_on_mobile": row_data.get("show on mobile") if "show on mobile" in row_data else row_data.get("show_on_mobile"),
            "show_on_desktop": row_data.get("show on desktop") if "show on desktop" in row_data else row_data.get("show_on_desktop"),
        }
        for field_name, value in bool_map.items():
            parsed = _coerce_bool(value)
            if parsed is not None:
                setattr(banner, field_name, parsed)

        banner.save()
        return "created" if created else "updated"

# ======================
# CUSTOM ADMIN SITE
# ======================

admin_site.register(Category, CategoryAdmin)
admin_site.register(Book, BookAdmin)
admin_site.register(Bundle, BundleAdmin)
admin_site.register(Subject, SubjectAdmin)
admin_site.register(Cart, CartAdmin)
admin_site.register(Wishlist, WishlistAdmin)
admin_site.register(Review, ReviewAdmin)
admin_site.register(Coupon, CouponAdmin)
admin_site.register(SyllabusPDF, SyllabusPDFAdmin)
admin_site.register(GalleryItem, GalleryItemAdmin)
admin_site.register(UserProfile, UserProfileAdmin)
admin_site.register(UserAddress, UserAddressAdmin)
admin_site.register(Order, OrderAdmin)
admin_site.register(PublishWithUsSubmission, PublishWithUsSubmissionAdmin)
admin_site.register(Banner, BannerAdmin)
admin_site.register(SearchQueryLog, SearchQueryLogAdmin)
admin_site.register(UnaniReferenceSource, UnaniReferenceSourceAdmin)
admin_site.register(DictionaryQueryLog, DictionaryQueryLogAdmin)
admin_site.register(DictionaryTermOpenLog, DictionaryTermOpenLogAdmin)
admin_site.register(AuditLog, AuditLogAdmin)
admin_site.register(SiteSettings, SiteSettingsAdmin)
admin_site.register(IKSCoinsSettings, IKSCoinsSettingsAdmin)
admin_site.register(IKSWallet, IKSWalletAdmin)
admin_site.register(IKSWalletTransaction, IKSWalletTransactionAdmin)
admin_site.register(UnaniTerm, UnaniTermAdmin)
admin_site.register(ClassicalWeightUnit, ClassicalWeightUnitAdmin)

