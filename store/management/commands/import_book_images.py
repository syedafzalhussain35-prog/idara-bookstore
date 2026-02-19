import re
import csv
import zipfile
import unicodedata
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlsplit
from urllib.request import Request, urlopen

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from store.models import Book, BookImage


FOLDER_PATTERN = re.compile(r"^(.*)\((.*)\)$")
VALID_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


def _image_sort_key(file_name):
    stem = file_name.rsplit(".", 1)[0]
    return (not stem.isdigit(), int(stem) if stem.isdigit() else stem.lower())


def _parse_folder_name(folder_name):
    match = FOLDER_PATTERN.match(folder_name or "")
    if not match:
        return None, None
    title = (match.group(1) or "").strip()
    author = (match.group(2) or "").strip()
    if not title or not author:
        return None, None
    return title, author


def _normalize_text(value):
    value = unicodedata.normalize("NFKC", str(value or ""))
    value = value.casefold()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _looks_like_url(value):
    text = str(value or "").strip().lower()
    return text.startswith("http://") or text.startswith("https://")


def _split_gallery_sources(raw):
    text = str(raw or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;\n|]+", text) if part.strip()]


class Command(BaseCommand):
    help = (
        "Bulk import book images from folders named 'Title (Author)' where "
        "1.jpg is main cover and the rest become gallery images."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--root-path",
            type=str,
            default="",
            help="Server path containing folders named 'Title (Author)'.",
        )
        parser.add_argument(
            "--zip-path",
            type=str,
            default="",
            help="Path to ZIP archive containing folders named 'Title (Author)'.",
        )
        parser.add_argument(
            "--excel-path",
            type=str,
            default="",
            help=(
                "Path to XLSX mapping file. Required columns: title, author (or system_id). "
                "Optional: main_cover / main_cover_url / main_cover_path, "
                "gallery / gallery_urls / gallery_paths, gallery_path_1..N, system_id."
            ),
        )
        parser.add_argument(
            "--excel-fallback-root",
            type=str,
            default="",
            help=(
                "Optional folder root for --excel-path fallback. If a row has no image paths, "
                "the command will try '<root>/Title (Author)'. Defaults to '<excel_dir>/bulk_images' if present."
            ),
        )
        parser.add_argument(
            "--report-path",
            type=str,
            default="",
            help=(
                "Optional CSV output path for skipped rows report. "
                "If omitted for --excel-path, defaults to '<excel_name>_import_skipped.csv'."
            ),
        )
        parser.add_argument(
            "--replace-existing",
            action="store_true",
            help="Delete existing main cover and gallery images before importing new ones.",
        )
        parser.add_argument(
            "--create-missing-books",
            action="store_true",
            help="When importing from --excel-path, create missing books using title/author/system_id.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and report without writing DB/files.",
        )
        parser.add_argument(
            "--fuzzy",
            action="store_true",
            help="Allow fuzzy Title/Author matching when exact match is not found.",
        )
        parser.add_argument(
            "--fuzzy-threshold",
            type=float,
            default=0.82,
            help="Minimum fuzzy match score between 0 and 1 (default: 0.82).",
        )

    def handle(self, *args, **options):
        root_path = (options.get("root_path") or "").strip()
        zip_path = (options.get("zip_path") or "").strip()
        excel_path = (options.get("excel_path") or "").strip()
        excel_fallback_root = (options.get("excel_fallback_root") or "").strip()
        report_path = (options.get("report_path") or "").strip()
        replace_existing = bool(options.get("replace_existing"))
        create_missing_books = bool(options.get("create_missing_books"))
        dry_run = bool(options.get("dry_run"))
        fuzzy = bool(options.get("fuzzy"))
        fuzzy_threshold = float(options.get("fuzzy_threshold"))

        selected_inputs = [bool(root_path), bool(zip_path), bool(excel_path)]
        if sum(selected_inputs) != 1:
            raise CommandError("Provide exactly one of --root-path, --zip-path, or --excel-path.")
        if not 0 <= fuzzy_threshold <= 1:
            raise CommandError("--fuzzy-threshold must be between 0 and 1.")

        book_index = self._build_book_index()
        skipped_details = []

        if root_path:
            created, skipped, errors, error_messages, skipped_details = self._import_from_root(
                Path(root_path),
                book_index=book_index,
                replace_existing=replace_existing,
                dry_run=dry_run,
                fuzzy=fuzzy,
                fuzzy_threshold=fuzzy_threshold,
            )
        elif zip_path:
            created, skipped, errors, error_messages, skipped_details = self._import_from_zip_path(
                Path(zip_path),
                book_index=book_index,
                replace_existing=replace_existing,
                dry_run=dry_run,
                fuzzy=fuzzy,
                fuzzy_threshold=fuzzy_threshold,
            )
        else:
            created, skipped, errors, error_messages, skipped_details = self._import_from_excel_path(
                Path(excel_path),
                book_index=book_index,
                replace_existing=replace_existing,
                create_missing_books=create_missing_books,
                excel_fallback_root=excel_fallback_root,
                dry_run=dry_run,
                fuzzy=fuzzy,
                fuzzy_threshold=fuzzy_threshold,
            )
            if skipped_details:
                if report_path:
                    report_file = Path(report_path)
                    if not report_file.is_absolute():
                        report_file = (Path(excel_path).parent / report_file).resolve()
                else:
                    report_file = Path(excel_path).with_name(
                        f"{Path(excel_path).stem}_import_skipped.csv"
                    )
                self._write_skipped_report(report_file, skipped_details)
                self.stdout.write(self.style.WARNING(f"Skipped rows report: {report_file}"))

        summary = (
            f"Import completed. Updated: {created}, Skipped: {skipped}, Errors: {errors}"
            + (" (dry run)" if dry_run else "")
        )
        if errors:
            self.stdout.write(self.style.WARNING(summary))
        else:
            self.stdout.write(self.style.SUCCESS(summary))

        for msg in error_messages[:50]:
            self.stdout.write(self.style.ERROR(msg))
        if len(error_messages) > 50:
            self.stdout.write(
                self.style.ERROR(f"Additional errors not shown: {len(error_messages) - 50}")
            )

    def _write_skipped_report(self, report_file, skipped_details):
        report_file.parent.mkdir(parents=True, exist_ok=True)
        with report_file.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(
                fh,
                fieldnames=[
                    "row_number",
                    "reason",
                    "title",
                    "author",
                    "system_id",
                    "main_source",
                    "gallery_sources",
                    "note",
                ],
            )
            writer.writeheader()
            for item in skipped_details:
                writer.writerow(item)

    def _build_book_index(self):
        exact = {}
        by_system_id = {}
        normalized = {}
        corpus = []
        for book in Book.objects.only("id", "title", "author", "system_id"):
            raw_key = ((book.title or "").strip(), (book.author or "").strip())
            exact[raw_key] = book
            if book.system_id:
                by_system_id[str(book.system_id).strip().casefold()] = book

            norm_title = _normalize_text(raw_key[0])
            norm_author = _normalize_text(raw_key[1])
            norm_key = (norm_title, norm_author)
            normalized.setdefault(norm_key, []).append(book)

            corpus.append((book, norm_title, norm_author))
        return {
            "exact": exact,
            "by_system_id": by_system_id,
            "normalized": normalized,
            "corpus": corpus,
        }

    def _resolve_book(self, title, author, book_index, fuzzy=False, fuzzy_threshold=0.82, system_id=""):
        sid = str(system_id or "").strip().casefold()
        if sid:
            by_sid = book_index["by_system_id"].get(sid)
            if by_sid:
                return by_sid

        raw_key = ((title or "").strip(), (author or "").strip())
        book = book_index["exact"].get(raw_key)
        if book:
            return book

        norm_title = _normalize_text(raw_key[0])
        norm_author = _normalize_text(raw_key[1])
        normalized_hits = book_index["normalized"].get((norm_title, norm_author), [])
        if len(normalized_hits) == 1:
            return normalized_hits[0]
        if len(normalized_hits) > 1:
            return None
        if not fuzzy:
            return None

        scores = []
        for candidate_book, cand_title, cand_author in book_index["corpus"]:
            title_score = SequenceMatcher(None, norm_title, cand_title).ratio()
            author_score = SequenceMatcher(None, norm_author, cand_author).ratio()
            overall = (0.72 * title_score) + (0.28 * author_score)
            scores.append((overall, candidate_book))

        if not scores:
            return None
        scores.sort(key=lambda item: item[0], reverse=True)
        best_score, best_book = scores[0]
        second_score = scores[1][0] if len(scores) > 1 else 0.0

        # Avoid risky auto-matches when top two candidates are too close.
        if best_score >= fuzzy_threshold and (best_score - second_score) >= 0.03:
            return best_book
        return None

    def _import_from_root(self, root, book_index, replace_existing=False, dry_run=False, fuzzy=False, fuzzy_threshold=0.82):
        if not root.exists() or not root.is_dir():
            raise CommandError(f"Root path does not exist or is not a directory: {root}")

        created = 0
        skipped = 0
        errors = 0
        error_messages = []

        for folder in sorted((p for p in root.iterdir() if p.is_dir()), key=lambda p: p.name.lower()):
            title, author = _parse_folder_name(folder.name)
            if not title or not author:
                skipped += 1
                continue

            book = self._resolve_book(
                title,
                author,
                book_index=book_index,
                fuzzy=fuzzy,
                fuzzy_threshold=fuzzy_threshold,
            )
            if not book:
                skipped += 1
                continue

            images = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in VALID_EXTENSIONS]
            images.sort(key=lambda p: _image_sort_key(p.name))
            if not images:
                skipped += 1
                continue

            try:
                with transaction.atomic():
                    if replace_existing and not dry_run:
                        if book.main_cover:
                            book.main_cover.delete(save=False)
                        book.images.all().delete()

                    if not dry_run:
                        main_image = images[0]
                        with main_image.open("rb") as fh:
                            book.main_cover.save(main_image.name, ContentFile(fh.read(), name=main_image.name), save=True)

                        for extra_image in images[1:]:
                            with extra_image.open("rb") as fh:
                                BookImage.objects.create(
                                    book=book,
                                    image=ContentFile(fh.read(), name=extra_image.name),
                                )

                created += 1
            except Exception as exc:
                errors += 1
                error_messages.append(f"{folder.name}: {exc}")

        return created, skipped, errors, error_messages, []

    def _import_from_zip_path(self, zip_path, book_index, replace_existing=False, dry_run=False, fuzzy=False, fuzzy_threshold=0.82):
        if not zip_path.exists() or not zip_path.is_file():
            raise CommandError(f"ZIP path does not exist or is not a file: {zip_path}")

        created = 0
        skipped = 0
        errors = 0
        error_messages = []

        grouped = {}
        try:
            with zipfile.ZipFile(zip_path) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue

                    path = str(info.filename or "").replace("\\", "/").strip("/")
                    if not path:
                        continue

                    parts = [part for part in path.split("/") if part]
                    if len(parts) < 2:
                        continue

                    file_name = parts[-1]
                    suffix = f".{file_name.rsplit('.', 1)[-1].lower()}" if "." in file_name else ""
                    if suffix not in VALID_EXTENSIONS:
                        continue

                    folder_name = None
                    for segment in reversed(parts[:-1]):
                        if FOLDER_PATTERN.match(segment):
                            folder_name = segment
                            break
                    if not folder_name:
                        continue

                    grouped.setdefault(folder_name, []).append((file_name, info))

                for folder_name in sorted(grouped.keys(), key=lambda x: x.lower()):
                    title, author = _parse_folder_name(folder_name)
                    if not title or not author:
                        skipped += 1
                        continue

                    book = self._resolve_book(
                        title,
                        author,
                        book_index=book_index,
                        fuzzy=fuzzy,
                        fuzzy_threshold=fuzzy_threshold,
                    )
                    if not book:
                        skipped += 1
                        continue

                    image_items = sorted(grouped.get(folder_name, []), key=lambda item: _image_sort_key(item[0]))
                    if not image_items:
                        skipped += 1
                        continue

                    try:
                        with transaction.atomic():
                            if replace_existing and not dry_run:
                                if book.main_cover:
                                    book.main_cover.delete(save=False)
                                book.images.all().delete()

                            if not dry_run:
                                main_name, main_info = image_items[0]
                                main_data = zf.read(main_info)
                                book.main_cover.save(main_name, ContentFile(main_data, name=main_name), save=True)

                                for extra_name, extra_info in image_items[1:]:
                                    extra_data = zf.read(extra_info)
                                    BookImage.objects.create(
                                        book=book,
                                        image=ContentFile(extra_data, name=extra_name),
                                    )

                        created += 1
                    except Exception as exc:
                        errors += 1
                        error_messages.append(f"{folder_name}: {exc}")
        except zipfile.BadZipFile as exc:
            raise CommandError(f"Invalid ZIP file: {exc}") from exc

        return created, skipped, errors, error_messages, []

    def _resolve_source(self, source, base_dir):
        source = str(source or "").strip()
        if not source:
            raise ValueError("empty image source")

        if _looks_like_url(source):
            request = Request(source, headers={"User-Agent": "IdaraBookstoreImporter/1.0"})
            with urlopen(request, timeout=30) as response:
                data = response.read()
            url_path = urlsplit(source).path
            file_name = Path(url_path).name or "image.jpg"
            return file_name, data

        source_path = Path(source)
        if not source_path.is_absolute():
            source_path = (base_dir / source_path).resolve()
        if not source_path.exists() or not source_path.is_file():
            raise FileNotFoundError(f"file not found: {source_path}")
        return source_path.name, source_path.read_bytes()

    def _import_from_excel_path(
        self,
        excel_path,
        book_index,
        replace_existing=False,
        create_missing_books=False,
        excel_fallback_root="",
        dry_run=False,
        fuzzy=False,
        fuzzy_threshold=0.82,
    ):
        if not excel_path.exists() or not excel_path.is_file():
            raise CommandError(f"Excel path does not exist or is not a file: {excel_path}")
        if excel_path.suffix.lower() not in {".xlsx"}:
            raise CommandError("Only .xlsx is supported for --excel-path.")

        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise CommandError("openpyxl is required for --excel-path.") from exc

        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return 0, 0, 0, ["Excel file is empty."]

        headers = [_normalize_header(item) for item in rows[0]]
        has_title_author = "title" in headers and "author" in headers
        has_system_id = "system_id" in headers
        if not has_title_author and not has_system_id:
            raise CommandError("Excel must include either system_id or both title and author columns.")

        created = 0
        skipped = 0
        errors = 0
        error_messages = []
        skipped_details = []
        base_dir = excel_path.parent
        fallback_root = None
        if str(excel_fallback_root or "").strip():
            configured_fallback = Path(excel_fallback_root)
            if not configured_fallback.is_absolute():
                configured_fallback = (base_dir / configured_fallback).resolve()
            if configured_fallback.exists() and configured_fallback.is_dir():
                fallback_root = configured_fallback
        else:
            default_fallback = (base_dir / "bulk_images").resolve()
            if default_fallback.exists() and default_fallback.is_dir():
                fallback_root = default_fallback
        fallback_index = {}
        if fallback_root:
            for folder in (p for p in fallback_root.iterdir() if p.is_dir()):
                f_title, f_author = _parse_folder_name(folder.name)
                if not f_title or not f_author:
                    continue
                key = (_normalize_text(f_title), _normalize_text(f_author))
                fallback_index.setdefault(key, []).append(folder)

        for row_number, row in enumerate(rows[1:], start=2):
            row_values = list(row or [])
            row_dict = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                row_dict[header] = row_values[idx] if idx < len(row_values) else None

            title = str(row_dict.get("title") or "").strip()
            author = str(row_dict.get("author") or "").strip()
            system_id = str(row_dict.get("system_id") or "").strip()
            if (not system_id) and (not title or not author):
                skipped += 1
                skipped_details.append(
                    {
                        "row_number": row_number,
                        "reason": "missing_identifier",
                        "title": title,
                        "author": author,
                        "system_id": system_id,
                        "main_source": "",
                        "gallery_sources": "",
                        "note": "Row requires system_id or both title and author.",
                    }
                )
                continue

            main_source = (
                row_dict.get("main_cover")
                or row_dict.get("main_cover_url")
                or row_dict.get("main_cover_path")
                or ""
            )
            gallery_sources = []
            gallery_headers = {
                "gallery",
                "gallery_url",
                "gallery_urls",
                "gallery_path",
                "gallery_paths",
            }
            for idx, header in enumerate(headers):
                if not header:
                    continue
                if header in gallery_headers or header.startswith("gallery_path_") or header.startswith("gallery_url_"):
                    raw_value = row_values[idx] if idx < len(row_values) else None
                    gallery_sources.extend(_split_gallery_sources(raw_value))

            if not gallery_sources:
                gallery_raw = (
                    row_dict.get("gallery")
                    or row_dict.get("gallery_urls")
                    or row_dict.get("gallery_paths")
                    or row_dict.get("gallery_path")
                    or row_dict.get("gallery_url")
                    or ""
                )
                gallery_sources = _split_gallery_sources(gallery_raw)

            # Preserve order and remove duplicates.
            unique_gallery_sources = []
            seen = set()
            for src in gallery_sources:
                key = src.strip()
                if not key:
                    continue
                if key in seen:
                    continue
                seen.add(key)
                unique_gallery_sources.append(key)
            gallery_sources = unique_gallery_sources

            if (
                fallback_root
                and title
                and author
                and not str(main_source or "").strip()
                and not gallery_sources
            ):
                folder = fallback_root / f"{title} ({author})"
                if not (folder.exists() and folder.is_dir()):
                    norm_key = (_normalize_text(title), _normalize_text(author))
                    matches = fallback_index.get(norm_key, [])
                    if len(matches) == 1:
                        folder = matches[0]
                    else:
                        folder = None
                if folder and folder.exists() and folder.is_dir():
                    files = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in VALID_EXTENSIONS]
                    files.sort(key=lambda p: _image_sort_key(p.name))
                    if files:
                        main_source = str(files[0])
                        gallery_sources = [str(p) for p in files[1:]]

            if not str(main_source or "").strip() and not gallery_sources:
                skipped += 1
                skipped_details.append(
                    {
                        "row_number": row_number,
                        "reason": "missing_sources",
                        "title": title,
                        "author": author,
                        "system_id": system_id,
                        "main_source": "",
                        "gallery_sources": "",
                        "note": "No main or gallery image source found in row/fallback.",
                    }
                )
                continue

            book = self._resolve_book(
                title,
                author,
                book_index=book_index,
                fuzzy=fuzzy,
                fuzzy_threshold=fuzzy_threshold,
                system_id=system_id,
            )

            if not book and create_missing_books and title and author:
                if not dry_run:
                    sid_value = system_id or None
                    if sid_value:
                        book, _ = Book.objects.get_or_create(
                            system_id=sid_value,
                            defaults={
                                "title": title,
                                "author": author,
                                "price": Decimal("0.00"),
                            },
                        )
                    else:
                        book, _ = Book.objects.get_or_create(
                            title=title,
                            author=author,
                            defaults={
                                "system_id": None,
                                "price": Decimal("0.00"),
                            },
                        )
                    raw_key = ((book.title or "").strip(), (book.author or "").strip())
                    book_index["exact"][raw_key] = book
                    if book.system_id:
                        book_index["by_system_id"][str(book.system_id).strip().casefold()] = book
                    norm_title = _normalize_text(raw_key[0])
                    norm_author = _normalize_text(raw_key[1])
                    book_index["normalized"].setdefault((norm_title, norm_author), []).append(book)
                    book_index["corpus"].append((book, norm_title, norm_author))
                else:
                    book = Book(
                        title=title,
                        author=author,
                        system_id=system_id or None,
                        price=Decimal("0.00"),
                    )

            if not book:
                skipped += 1
                skipped_details.append(
                    {
                        "row_number": row_number,
                        "reason": "book_not_found",
                        "title": title,
                        "author": author,
                        "system_id": system_id,
                        "main_source": str(main_source or "").strip(),
                        "gallery_sources": "; ".join(gallery_sources),
                        "note": "No matching book found.",
                    }
                )
                continue

            try:
                with transaction.atomic():
                    if replace_existing and not dry_run:
                        if book.main_cover:
                            book.main_cover.delete(save=False)
                        book.images.all().delete()

                    if not dry_run:
                        if str(main_source or "").strip():
                            main_name, main_data = self._resolve_source(main_source, base_dir)
                            book.main_cover.save(main_name, ContentFile(main_data, name=main_name), save=True)

                        for source in gallery_sources:
                            image_name, image_data = self._resolve_source(source, base_dir)
                            BookImage.objects.create(
                                book=book,
                                image=ContentFile(image_data, name=image_name),
                            )

                created += 1
            except Exception as exc:
                errors += 1
                error_messages.append(f"Row {row_number}: {exc}")

        return created, skipped, errors, error_messages, skipped_details
